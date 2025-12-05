"""
SSH Guardian v3.0 - Export Routes
API endpoints for exporting data in CSV, JSON, and XLSX formats
"""

import sys
import io
import csv
import json
from pathlib import Path
from datetime import datetime, timedelta
from decimal import Decimal
from flask import Blueprint, jsonify, request, Response

# Add project paths
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
sys.path.append(str(PROJECT_ROOT / "dbs"))
sys.path.append(str(PROJECT_ROOT / "src"))

from connection import get_connection

# Create Blueprint
export_routes = Blueprint('export_routes', __name__, url_prefix='/api/dashboard/export')


def get_date_range(date_range: str, start_date: str = None, end_date: str = None):
    """Convert date range parameter to start/end dates"""
    today = datetime.now().date()

    if date_range == 'custom' and start_date and end_date:
        return start_date, end_date
    elif date_range == 'today':
        return str(today), str(today)
    elif date_range == '7days':
        return str(today - timedelta(days=7)), str(today)
    elif date_range == '30days':
        return str(today - timedelta(days=30)), str(today)
    elif date_range == '90days':
        return str(today - timedelta(days=90)), str(today)
    else:
        # Default to last 7 days
        return str(today - timedelta(days=7)), str(today)


def format_value(value):
    """Format a value for export"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bytes):
        # Handle binary IP addresses
        return value.hex()
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    return str(value)


def generate_csv(data: list, filename: str) -> Response:
    """Generate CSV response from data"""
    if not data:
        output = io.StringIO()
        output.write('No data available\n')
        response = Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
        return response

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()

    for row in data:
        formatted_row = {k: format_value(v) for k, v in row.items()}
        writer.writerow(formatted_row)

    response = Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )
    return response


def generate_json_export(data: list, filename: str) -> Response:
    """Generate JSON response from data"""
    # Format datetime objects and other non-serializable types
    formatted_data = []
    for row in data:
        formatted_row = {}
        for k, v in row.items():
            if isinstance(v, datetime):
                formatted_row[k] = v.isoformat()
            elif isinstance(v, Decimal):
                formatted_row[k] = float(v)
            elif isinstance(v, bytes):
                formatted_row[k] = v.hex()
            else:
                formatted_row[k] = v
        formatted_data.append(formatted_row)

    response = Response(
        json.dumps({'data': formatted_data, 'count': len(formatted_data)}, indent=2, default=str),
        mimetype='application/json',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )
    return response


def generate_xlsx(data: list, filename: str) -> Response:
    """Generate XLSX response from data"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        if not data:
            ws['A1'] = 'No data available'
        else:
            # Header row with styling
            headers = list(data[0].keys())
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Data rows
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header)
                    ws.cell(row=row_idx, column=col_idx, value=format_value(value))

            # Auto-width columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
        return response

    except ImportError:
        # openpyxl not installed, fall back to CSV
        return generate_csv(data, filename.replace('.xlsx', '.csv'))


@export_routes.route('/events', methods=['GET'])
def export_events():
    """Export authentication events"""
    try:
        format_type = request.args.get('format', 'csv')
        limit = min(int(request.args.get('limit', 10000)), 1000000)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        where_clauses = []
        params = []

        if start_date:
            where_clauses.append("DATE(ae.timestamp) >= %s")
            params.append(start_date)

        if end_date:
            where_clauses.append("DATE(ae.timestamp) <= %s")
            params.append(end_date)

        where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        query = f"""
            SELECT
                ae.id, ae.source_ip_text as source_ip, ae.target_username as username,
                ae.event_type, ae.timestamp, ae.target_server as server_name,
                ae.target_port as port, ae.auth_method, ae.failure_reason,
                ae.ml_risk_score, ae.ml_threat_type, ae.is_anomaly,
                ae.processing_status, ae.source_type
            FROM auth_events ae
            {where_sql}
            ORDER BY ae.timestamp DESC
            LIMIT %s
        """
        params.append(limit)

        cursor.execute(query, params)
        data = cursor.fetchall()

        cursor.close()
        conn.close()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"events_export_{timestamp}.{format_type}"

        if format_type == 'json':
            return generate_json_export(data, filename)
        elif format_type == 'xlsx':
            return generate_xlsx(data, filename)
        else:
            return generate_csv(data, filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@export_routes.route('/blocked_ips', methods=['GET'])
def export_blocked_ips():
    """Export blocked IPs"""
    try:
        format_type = request.args.get('format', 'csv')
        limit = min(int(request.args.get('limit', 10000)), 1000000)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        where_clauses = []
        params = []

        if start_date:
            where_clauses.append("DATE(blocked_at) >= %s")
            params.append(start_date)

        if end_date:
            where_clauses.append("DATE(blocked_at) <= %s")
            params.append(end_date)

        where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        query = f"""
            SELECT
                id, ip_address_text as source_ip, block_reason, block_source,
                blocked_at, unblock_at, auto_unblock, is_active,
                failed_attempts, risk_score, threat_level,
                is_simulation, created_at
            FROM ip_blocks
            {where_sql}
            ORDER BY blocked_at DESC
            LIMIT %s
        """
        params.append(limit)

        cursor.execute(query, params)
        data = cursor.fetchall()

        cursor.close()
        conn.close()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"blocked_ips_export_{timestamp}.{format_type}"

        if format_type == 'json':
            return generate_json_export(data, filename)
        elif format_type == 'xlsx':
            return generate_xlsx(data, filename)
        else:
            return generate_csv(data, filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@export_routes.route('/ip_stats', methods=['GET'])
def export_ip_stats():
    """Export IP statistics"""
    try:
        format_type = request.args.get('format', 'csv')
        limit = min(int(request.args.get('limit', 10000)), 1000000)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        where_clauses = []
        params = []

        if start_date:
            where_clauses.append("DATE(last_seen) >= %s")
            params.append(start_date)

        if end_date:
            where_clauses.append("DATE(last_seen) <= %s")
            params.append(end_date)

        where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        query = f"""
            SELECT
                id, ip_address_text as source_ip, total_events, failed_events,
                successful_events, invalid_events, unique_servers, unique_usernames,
                avg_risk_score, max_risk_score, anomaly_count,
                times_blocked, currently_blocked, first_seen, last_seen
            FROM ip_statistics
            {where_sql}
            ORDER BY total_events DESC
            LIMIT %s
        """
        params.append(limit)

        cursor.execute(query, params)
        data = cursor.fetchall()

        cursor.close()
        conn.close()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"ip_stats_export_{timestamp}.{format_type}"

        if format_type == 'json':
            return generate_json_export(data, filename)
        elif format_type == 'xlsx':
            return generate_xlsx(data, filename)
        else:
            return generate_csv(data, filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@export_routes.route('/threat_intel', methods=['GET'])
def export_threat_intel():
    """Export threat intelligence data"""
    try:
        format_type = request.args.get('format', 'csv')
        limit = min(int(request.args.get('limit', 10000)), 1000000)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        where_clauses = []
        params = []

        if start_date:
            where_clauses.append("DATE(updated_at) >= %s")
            params.append(start_date)

        if end_date:
            where_clauses.append("DATE(updated_at) <= %s")
            params.append(end_date)

        where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        query = f"""
            SELECT
                id, ip_address_text as source_ip,
                abuseipdb_score, abuseipdb_confidence, abuseipdb_reports,
                abuseipdb_checked_at,
                virustotal_positives, virustotal_total, virustotal_checked_at,
                overall_threat_level, threat_confidence,
                created_at, updated_at
            FROM ip_threat_intelligence
            {where_sql}
            ORDER BY updated_at DESC
            LIMIT %s
        """
        params.append(limit)

        cursor.execute(query, params)
        data = cursor.fetchall()

        cursor.close()
        conn.close()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"threat_intel_export_{timestamp}.{format_type}"

        if format_type == 'json':
            return generate_json_export(data, filename)
        elif format_type == 'xlsx':
            return generate_xlsx(data, filename)
        else:
            return generate_csv(data, filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@export_routes.route('/geoip', methods=['GET'])
def export_geoip():
    """Export GeoIP data"""
    try:
        format_type = request.args.get('format', 'csv')
        limit = min(int(request.args.get('limit', 10000)), 1000000)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        where_clauses = []
        params = []

        if start_date:
            where_clauses.append("DATE(last_seen) >= %s")
            params.append(start_date)

        if end_date:
            where_clauses.append("DATE(last_seen) <= %s")
            params.append(end_date)

        where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        query = f"""
            SELECT
                id, ip_address_text as source_ip, country_code, country_name,
                region, city, latitude, longitude, timezone,
                asn, asn_org, isp,
                is_proxy, is_vpn, is_tor, is_datacenter,
                first_seen, last_seen
            FROM ip_geolocation
            {where_sql}
            ORDER BY last_seen DESC
            LIMIT %s
        """
        params.append(limit)

        cursor.execute(query, params)
        data = cursor.fetchall()

        cursor.close()
        conn.close()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"geoip_export_{timestamp}.{format_type}"

        if format_type == 'json':
            return generate_json_export(data, filename)
        elif format_type == 'xlsx':
            return generate_xlsx(data, filename)
        else:
            return generate_csv(data, filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@export_routes.route('/audit', methods=['GET'])
def export_audit():
    """Export audit logs"""
    try:
        format_type = request.args.get('format', 'csv')
        limit = min(int(request.args.get('limit', 10000)), 1000000)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        where_clauses = []
        params = []

        if start_date:
            where_clauses.append("DATE(a.created_at) >= %s")
            params.append(start_date)

        if end_date:
            where_clauses.append("DATE(a.created_at) <= %s")
            params.append(end_date)

        where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        query = f"""
            SELECT
                a.id, a.user_id, a.action, a.resource_type, a.resource_id,
                a.details, a.ip_address, a.user_agent, a.created_at,
                u.email as user_email, u.full_name as user_name
            FROM audit_logs a
            LEFT JOIN users u ON a.user_id = u.id
            {where_sql}
            ORDER BY a.created_at DESC
            LIMIT %s
        """
        params.append(limit)

        cursor.execute(query, params)
        data = cursor.fetchall()

        cursor.close()
        conn.close()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"audit_export_{timestamp}.{format_type}"

        if format_type == 'json':
            return generate_json_export(data, filename)
        elif format_type == 'xlsx':
            return generate_xlsx(data, filename)
        else:
            return generate_csv(data, filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@export_routes.route('/notifications', methods=['GET'])
def export_notifications():
    """Export notification history"""
    try:
        format_type = request.args.get('format', 'csv')
        limit = min(int(request.args.get('limit', 10000)), 1000000)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        where_clauses = []
        params = []

        if start_date:
            where_clauses.append("DATE(n.created_at) >= %s")
            params.append(start_date)

        if end_date:
            where_clauses.append("DATE(n.created_at) <= %s")
            params.append(end_date)

        where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        query = f"""
            SELECT
                n.id, n.notification_rule_id, n.trigger_type,
                n.trigger_event_id, n.trigger_block_id,
                n.message_title, n.message_body, n.priority,
                n.status, n.sent_at, n.failed_reason,
                n.retry_count, n.delivery_status, n.created_at,
                nr.name as rule_name
            FROM notifications n
            LEFT JOIN notification_rules nr ON n.notification_rule_id = nr.id
            {where_sql}
            ORDER BY n.created_at DESC
            LIMIT %s
        """
        params.append(limit)

        cursor.execute(query, params)
        data = cursor.fetchall()

        cursor.close()
        conn.close()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"notifications_export_{timestamp}.{format_type}"

        if format_type == 'json':
            return generate_json_export(data, filename)
        elif format_type == 'xlsx':
            return generate_xlsx(data, filename)
        else:
            return generate_csv(data, filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@export_routes.route('/ml_predictions', methods=['GET'])
def export_ml_predictions():
    """Export ML predictions"""
    try:
        format_type = request.args.get('format', 'csv')
        limit = min(int(request.args.get('limit', 10000)), 1000000)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        where_clauses = []
        params = []

        if start_date:
            where_clauses.append("DATE(mp.created_at) >= %s")
            params.append(start_date)

        if end_date:
            where_clauses.append("DATE(mp.created_at) <= %s")
            params.append(end_date)

        where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        query = f"""
            SELECT
                mp.id, mp.event_id, mp.model_id,
                mp.risk_score, mp.threat_type, mp.confidence,
                mp.is_anomaly, mp.inference_time_ms,
                mp.was_blocked, mp.manual_feedback,
                mp.created_at,
                mm.model_name, mm.algorithm
            FROM ml_predictions mp
            LEFT JOIN ml_models mm ON mp.model_id = mm.id
            {where_sql}
            ORDER BY mp.created_at DESC
            LIMIT %s
        """
        params.append(limit)

        cursor.execute(query, params)
        data = cursor.fetchall()

        cursor.close()
        conn.close()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"ml_predictions_export_{timestamp}.{format_type}"

        if format_type == 'json':
            return generate_json_export(data, filename)
        elif format_type == 'xlsx':
            return generate_xlsx(data, filename)
        else:
            return generate_csv(data, filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500
